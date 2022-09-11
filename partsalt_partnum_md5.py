import cupy as cp
import os
import time
from openpyxl import load_workbook

# Соль указывается допустимый набор, которое мы знаем
SALT = "49920647"
# Набор телефонного номера нужно указать начало и конец словаря
T_NUM = [89011110000, 89011119999]

md5_list = open('hashcat/md5_list.txt', "w")
wb = load_workbook("scoring_data_v.0.3.3.xlsx")
ws = wb["A2"]

# сравнения двух номеров телефона в GPU kernel
kernel = cp.ElementwiseKernel(
     'int64 t_num1, int64 t_num2, int64 arg_s1, int64 arg_s2 ', 'int64 salt',
     '''
     if ((arg_s1 - t_num1) == (arg_s2 - t_num2))
        {
               salt = arg_s1 - t_num1;
        }
        else{ 
               salt=0;
        }''')

# Набор обезличенных данных преобразуем виде текстового документа для быстрого перебора
for i in range(2,len(ws["A"])+1):
    md5_list.write(f"{ws['A' + str(i)].value}\n")
md5_list.close()

# Проверка текстового документа на наличия заполнения
read_md5_list = open("hashcat/md5_list.txt", "r")
print(f"len = {bool(read_md5_list)}")

# Вызываем библиотеку hashcat. Перебор данных осуществляется с помощью масок с длиной 11 символов
start_bf = time.time()
if bool(read_md5_list) is True:
    os.system("rm -rf ./hashcat/hashcat.potfile")
    os.system("./hashcat/hashcat.bin -m0 -a3 ./hashcat/md5_list.txt ?d?d?d?d?d?d?d?d?d?d?d")
end_bf = time.time()

# подключения к списку деобезличенного набора
potfile = open("hashcat/hashcat.potfile", "r")
array_salts = []
for md5 in potfile:
     array_salts.append(int(md5[33:44]))
sort_salts = sorted(array_salts)
cp_salts = cp.array(sort_salts)

# Ф-я для формирования набора солей
def result_salt(t_number,cp_salt):
     arg_s1 = cp.repeat(cp_salt[0], len(t_number))
     arg_s2 = cp.repeat(cp_salt[1], len(t_number))
     array_result = []
     for i in range(len(t_number)):
          arg_x = cp.repeat(t_number[i],len(t_number))
          arg_result = kernel(arg_x,t_number,arg_s1,arg_s2)
          arg_result = cp.trim_zeros(arg_result)
          if bool(arg_result) == True:
               array_result.append(list(cp.asnumpy(arg_result))[0])
     return array_result

# Ф-я сравнения набора солей с доступа с ограниченной искомой солью
def compare(salt,array_res):
     salts = []
     for arg in array_res:
          if salt[0:len(salt)] == str(arg)[0:len(salt)]:
               salts.append(arg)
     return salts
# Сравнения двух доступных массивов
def one_set(first,second):
     one_set = set(first) & set(second)
     return one_set

# Перебор возможных солей
def set_sum(t_num,cp_salt):

     res = result_salt(t_num, [cp_salt[0], cp_salt[1]])
     # res2 = result_salt(t_num, [cp_salt[2], cp_salt[3]])
     one = compare(SALT,res)
     for i in range(4,len(cp_salt)):
          if (i+1) > len(cp_salt) or len(one) < 75: break
          one = one_set(one,result_salt(t_num,[cp_salt[i], cp_salts[i+1]]))
          print(f"iteration[{i}] = {len(one)}")
     return one

# Набор телефонных номеров
t_num = cp.arange(T_NUM[0], T_NUM[1])
# array_result = (result_salt(t_num, cp_salts))
start_sm = time.time()
array_res = set_sum(t_num,cp_salts)
end_sm = time.time()
# a = set(compare(SALT,array_result)) & set(se)



print(array_res)
print(f"time for brute force:{end_bf-start_bf}")
print(f"time for salt:{end_sm-start_sm}")