import os
from openpyxl import load_workbook
import time

# Подключаемся к обезличенному набору данных
md5_list = open('hashcat/md5_list.txt', "w")
wb = load_workbook("scoring_data_v.0.3.3.xlsx")
ws = wb["A2"]
# Oтбор соли с существующими 3 номерами телефона
t_num = [89106881090,89057801765,89809229343]

# Набор обезличенных данных преобразуем виде текстова документа для быстрого перебора
for i in range(2,len(ws["A"])+1):
    md5_list.write(f"{ws['A' + str(i)].value}\n")
md5_list.close()

# Проверка текстового документа на наличия заполнения
read_md5_list = open("hashcat/md5_list.txt", "r")
print(f"len = {bool(read_md5_list)}")

# Вызываем библиотеку hashcat. Перебор данных осуществляется с помощью масок с длиной 11 символов
start_bf = time.time()
if bool(read_md5_list) is True:
    os.system("rm -rf ./hashcat/hashcat.potfile")
    os.system("./hashcat/hashcat.bin -m0 -a3 ./hashcat/md5_list.txt ?d?d?d?d?d?d?d?d?d?d?d")
end_bf = time.time()
# Oтбор соли с существующими 3 номерами телефона
potfile = open("hashcat/hashcat.potfile", "r")
dict_md5 = {}
array_salts_num = []

# Ф-я соли
def salts(t_num:int,array_salts_num):
    salt = []
    for salts_num in array_salts_num:
        salt.append(salts_num-t_num)
    return salt
# Создания словаря md5
for md5 in potfile:
    array_salts_num.append(int(md5[33:44]))
    dict_md5[str(md5[:32])] = md5[33:44]

# На примеры 3 номеров, выявляем истинную соль
start_salt = time.time()
salts1 = salts(t_num[0],array_salts_num)
salts2 = salts(t_num[1],array_salts_num)
salts3 = salts(t_num[2],array_salts_num)
salt_res = set(salts1) & set(salts2) & set(salts3)
end_salt = time.time()
print(f"salt[{len(salt_res)}] = {sorted(salt_res)}")

# Деобезличенный набор данных хранится в scoring_data_v.1.0.xlsx
for row in range(2,len(ws["A"])+1):
    if dict_md5[str(ws['A' + str(row)].value)]:
          ws['A' + str(row)] = int(dict_md5[str(ws['A' + str(row)].value)])-list(salt_res)[0]

ws['C1'] = "Соль"
ws['E2'] = f"Расчетное время по brute force = {(round(end_bf - start_bf,2))} сек"
ws['E3'] = f"Расчетное время по соли = {round(end_salt - start_salt,2)} сек"
print(f"Расчетное время по brute force = {round(end_bf - start_bf,2)} сек")
print(f"Расчетное время по соли = {round(end_salt - start_salt,2)} сек")

wb.save(filename="scoring_data_v.1.0.xlsx")

print("result can see in scoring_data_v.1.0.xlsx")
os.system("libreoffice --writer scoring_data_v.1.0.xlsx")

