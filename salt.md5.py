import os
from openpyxl import load_workbook
import time
import pandas as pd

# Подключаемся к обезличенному набору данных
md5_list = open('hashcat/md5_list.txt', "w")
wb = load_workbook("scoring_data_v.0.3.3.xlsx")
ws = wb["A2"]
SALT = 4992064728

# Набор обезличенных данных преобразуем виде текстова документа для быстрого перебора
for i in range(2,len(ws["A"])+1):
    md5_list.write(f"{ws['A' + str(i)].value}\n")
md5_list.close()

# Проверка текстового документа на наличия заполнения
read_md5_list = open("hashcat/md5_list.txt", "r")
print(f"len = {bool(read_md5_list)}")

# Вызываем библиотеку hashcat. Перебор данных осуществляется с помощью масок с длиной 11 символов
start_bf = time.time()
if bool(read_md5_list) is True:
    os.system("rm -rf ./hashcat/hashcat.potfile")
    os.system("./hashcat/hashcat.bin -m0 -a3 ./hashcat/md5_list.txt ?d?d?d?d?d?d?d?d?d?d?d")
end_bf = time.time()
# Oтбор соли с существующими 3 номерами телефона
potfile = open("hashcat/hashcat.potfile", "r")
dict_md5 = {}
array_salts_num = []

# Создания словаря md5
for md5 in potfile:
    array_salts_num.append(int(md5[33:44]))
    dict_md5[str(md5[:32])] = md5[33:44]

# Деобезличенный набор данных хранится в scoring_data_v.1.0.xlsx
for row in range(2,len(ws["A"])+1):
    if dict_md5[str(ws['A' + str(row)].value)]:
          ws['A' + str(row)] = int(dict_md5[str(ws['A' + str(row)].value)])-SALT
          ws['C' + str(row)] = SALT
ws['C1'] = "Соль"
ws['E2'] = f"Расчетное время по brute force = {(round(end_bf - start_bf,2))} сек"
print(f"Расчетное время по brute force = {round(end_bf - start_bf,2)} сек")

PATH_FILE_NAME = "scoring_data_v.1.0.xlsx"
wb.save(filename=PATH_FILE_NAME)

print("result can see in scoring_data_v.1.0.xlsx")
os.system("libreoffice --writer scoring_data_v.1.0.xlsx")
print(pd.read_excel(PATH_FILE_NAME,usecols=["Номер телефона","Скоринговый балл","Соль"]))
